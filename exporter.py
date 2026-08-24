import os
import csv
from datetime import datetime
from typing import List, Optional


def export_to_excel(
    links: List[str],
    output_path: Optional[str] = None,
    column_name: str = "link",
    export_dir: str = "exports",
) -> str:
    """
    Xuất danh sách link video ra file Excel (.xlsx).
    Nếu chưa cài thư viện openpyxl, tự động fallback về .csv (UTF-8 with BOM mở trực tiếp bằng Excel).

    Args:
        links: Danh sách đường dẫn video.
        output_path: Đường dẫn file xuất (nếu None sẽ tự động đặt tên theo ngày giờ).
        column_name: Tên tiêu đề cột (mặc định 'link').
        export_dir: Thư mục chứa file xuất nếu không chỉ định output_path.

    Returns:
        Đường dẫn file đã được tạo thành công.
    """
    os.makedirs(export_dir, exist_ok=True)

    if not output_path:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(export_dir, f"tiktok_links_{timestamp}.xlsx")

    # Đảm bảo thư mục cha của output_path tồn tại
    parent_dir = os.path.dirname(output_path)
    if parent_dir:
        os.makedirs(parent_dir, exist_ok=True)

    # Thử xuất định dạng .xlsx bằng openpyxl
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "TikTok Links"

        # Style cho Header
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # Ghi tiêu đề cột: mặc định |link|
        ws.cell(row=1, column=1, value=column_name)
        header_cell = ws.cell(row=1, column=1)
        header_cell.fill = header_fill
        header_cell.font = header_font
        header_cell.alignment = header_alignment
        ws.row_dimensions[1].height = 26

        # Ghi dữ liệu
        data_font = Font(name="Calibri", size=10, color="0563C1", underline="single")
        data_alignment = Alignment(horizontal="left", vertical="center")

        for idx, link in enumerate(links, start=2):
            cell = ws.cell(row=idx, column=1, value=link)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
            ws.row_dimensions[idx].height = 20

        # Auto fit độ rộng cột
        max_len = max([len(str(l)) for l in links] + [len(column_name)]) if links else len(column_name)
        ws.column_dimensions['A'].width = min(max(max_len + 4, 30), 90)

        wb.save(output_path)
        return output_path

    except ImportError:
        # Fallback về CSV nếu không có openpyxl
        if not output_path.endswith(".csv"):
            output_path = os.path.splitext(output_path)[0] + ".csv"

        with open(output_path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([column_name])
            for link in links:
                writer.writerow([link])

        return output_path
