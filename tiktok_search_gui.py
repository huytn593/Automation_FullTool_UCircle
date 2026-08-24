import os
import sys
import json
import threading
from datetime import datetime
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from tiktok_search_core import TikTokSearchEngine

class TikTokSearchApp:
    def __init__(self, root):
        self.root = root
        self.root.title("TikTok Video Search & Excel Exporter (Anti-Captcha)")
        self.root.geometry("990x690")
        self.root.minsize(890, 560)

        self.style = ttk.Style()
        self.style.theme_use("clam")
        
        self.bg_color = "#f8f9fa"
        self.primary_color = "#fe2c55" # TikTok Pink
        self.dark_color = "#161823"
        
        self.root.configure(bg=self.bg_color)
        self.data_results = []

        self._setup_ui()

    def _setup_ui(self):
        # Header
        header_frame = tk.Frame(self.root, bg=self.dark_color, height=75)
        header_frame.pack(fill="x", side="top")

        title_lbl = tk.Label(
            header_frame, 
            text="🎵 TIKTOK REAL VIDEO SEARCH & EXCEL EXPORTER", 
            font=("Segoe UI", 15, "bold"), 
            fg="white", 
            bg=self.dark_color
        )
        title_lbl.pack(pady=(12, 2))

        sub_lbl = tk.Label(
            header_frame, 
            text="Trích xuất Link Video TikTok THẬT 100% (@username/video/id) & Caption chuẩn -> Xuất Excel", 
            font=("Segoe UI", 9), 
            fg="#25f4ee", 
            bg=self.dark_color
        )
        sub_lbl.pack(pady=(0, 10))

        # Main container
        container = tk.Frame(self.root, bg=self.bg_color, padx=20, pady=12)
        container.pack(fill="both", expand=True)

        # Input Control Frame
        ctrl_frame = tk.LabelFrame(container, text=" Thiết lập tìm kiếm ", font=("Segoe UI", 10, "bold"), bg=self.bg_color, padx=12, pady=10)
        ctrl_frame.pack(fill="x", pady=(0, 10))

        # Keyword
        tk.Label(ctrl_frame, text="Từ khóa (Keyword):", font=("Segoe UI", 10), bg=self.bg_color).grid(row=0, column=0, sticky="w", padx=(0, 8), pady=5)
        self.txt_keyword = ttk.Entry(ctrl_frame, font=("Segoe UI", 10), width=30)
        self.txt_keyword.grid(row=0, column=1, sticky="w", padx=(0, 15), pady=5)
        self.txt_keyword.focus_set()
        self.txt_keyword.bind("<Return>", lambda e: self.start_search("auto"))

        # Limit
        tk.Label(ctrl_frame, text="Số lượng:", font=("Segoe UI", 10), bg=self.bg_color).grid(row=0, column=2, sticky="w", padx=(0, 8), pady=5)
        self.txt_limit = ttk.Spinbox(ctrl_frame, from_=1, to=100, font=("Segoe UI", 10), width=6)
        self.txt_limit.set(10)
        self.txt_limit.grid(row=0, column=3, sticky="w", padx=(0, 15), pady=5)

        # Fast Search Button (Không dính Captcha)
        self.btn_search_fast = tk.Button(
            ctrl_frame, 
            text="⚡ Tìm Nhanh (Không Captcha)", 
            font=("Segoe UI", 9, "bold"), 
            bg=self.primary_color, 
            fg="white", 
            activebackground="#e02449", 
            activeforeground="white", 
            padx=12, 
            pady=5, 
            cursor="hand2", 
            command=lambda: self.start_search("fast"),
            relief="flat"
        )
        self.btn_search_fast.grid(row=0, column=4, sticky="e", padx=(0, 8), pady=5)

        # Browser Search Button (Hiện cửa sổ tự giải Captcha nếu có)
        self.btn_search_browser = tk.Button(
            ctrl_frame, 
            text="🌐 Mở Trình duyệt Chrome", 
            font=("Segoe UI", 9, "bold"), 
            bg="#0d6efd", 
            fg="white", 
            activebackground="#0b5ed7", 
            activeforeground="white", 
            padx=12, 
            pady=5, 
            cursor="hand2", 
            command=lambda: self.start_search("browser"),
            relief="flat"
        )
        self.btn_search_browser.grid(row=0, column=5, sticky="e", padx=(0, 0), pady=5)

        # Status & Progress bar
        self.status_frame = tk.Frame(container, bg=self.bg_color)
        self.status_frame.pack(fill="x", pady=(0, 5))

        self.lbl_status = tk.Label(self.status_frame, text="Sẵn sàng quét video TikTok thật...", font=("Segoe UI", 9), fg="#555555", bg=self.bg_color)
        self.lbl_status.pack(side="left")

        self.progress = ttk.Progressbar(self.status_frame, mode="indeterminate", length=150)

        # Results Table
        table_frame = tk.Frame(container, bg=self.bg_color)
        table_frame.pack(fill="both", expand=True, pady=5)

        columns = ("stt", "link", "caption")
        self.tree = ttk.Treeview(table_frame, columns=columns, show="headings", selectmode="browse")
        
        self.tree.heading("stt", text="STT")
        self.tree.heading("link", text="Link Video TikTok Thật")
        self.tree.heading("caption", text="Caption / Mô tả")

        self.tree.column("stt", width=50, minwidth=40, anchor="center")
        self.tree.column("link", width=420, minwidth=280, anchor="w")
        self.tree.column("caption", width=430, minwidth=280, anchor="w")

        v_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        h_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        v_scroll.grid(row=0, column=1, sticky="ns")
        h_scroll.grid(row=1, column=0, sticky="ew")

        table_frame.grid_rowconfigure(0, weight=1)
        table_frame.grid_columnconfigure(0, weight=1)

        # Buttons Bottom
        btn_frame = tk.Frame(container, bg=self.bg_color, pady=10)
        btn_frame.pack(fill="x")

        self.lbl_count = tk.Label(btn_frame, text="Tổng số: 0 video", font=("Segoe UI", 10, "bold"), fg=self.dark_color, bg=self.bg_color)
        self.lbl_count.pack(side="left")

        self.btn_export = tk.Button(
            btn_frame, 
            text="📊 Xuất ra File Excel (.xlsx)", 
            font=("Segoe UI", 10, "bold"), 
            bg="#107c41", 
            fg="white", 
            activebackground="#0b582e", 
            activeforeground="white", 
            padx=18, 
            pady=6, 
            cursor="hand2", 
            command=self.export_to_excel,
            relief="flat",
            state="disabled"
        )
        self.btn_export.pack(side="right", padx=(10, 0))

        self.btn_clear = tk.Button(
            btn_frame, 
            text="🗑️ Xóa danh sách", 
            font=("Segoe UI", 9), 
            bg="#e9ecef", 
            fg="#495057", 
            padx=12, 
            pady=6, 
            cursor="hand2", 
            command=self.clear_results,
            relief="flat"
        )
        self.btn_clear.pack(side="right")

    def start_search(self, mode="auto"):
        keyword = self.txt_keyword.get().strip()
        if not keyword:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập từ khóa tìm kiếm!")
            return

        try:
            limit = int(self.txt_limit.get().strip())
            if limit <= 0:
                limit = 10
        except ValueError:
            limit = 10

        self.btn_search_fast.config(state="disabled")
        self.btn_search_browser.config(state="disabled")
        self.btn_export.config(state="disabled")
        
        msg = f"⏳ Đang quét video cho '{keyword}'..."
        if mode == "browser":
            msg = f"⏳ Đang mở Chrome... Hãy kéo Captcha nếu TikTok yêu cầu!"
        self.lbl_status.config(text=msg, fg="#0d6efd")
        
        self.progress.pack(side="right")
        self.progress.start(10)

        thread = threading.Thread(target=self._run_search_thread, args=(keyword, limit, mode), daemon=True)
        thread.start()

    def _run_search_thread(self, keyword, limit, mode):
        videos, method = TikTokSearchEngine.search(keyword, limit, mode)
        self.root.after(0, self._handle_search_result, videos, keyword, method)

    def _handle_search_result(self, results, keyword, method):
        self.progress.stop()
        self.progress.pack_forget()
        self.btn_search_fast.config(state="normal")
        self.btn_search_browser.config(state="normal")

        self.data_results = []
        for row in self.tree.get_children():
            self.tree.delete(row)

        for i, item in enumerate(results, 1):
            url = item.get("url", "")
            desc = item.get("desc", "")
            self.data_results.append({"stt": i, "url": url, "caption": desc})
            self.tree.insert("", "end", values=(i, url, desc))

        count = len(self.data_results)
        self.lbl_count.config(text=f"Tổng số: {count} video")
        
        if count > 0:
            self.lbl_status.config(text=f"✅ Tìm thấy {count} video TikTok thật ({method}).", fg="#198754")
            self.btn_export.config(state="normal")
        else:
            self.lbl_status.config(text=f"⚠️ Không tìm thấy video nào.", fg="#ffc107")
            messagebox.showinfo("Thông báo", f"Không tìm thấy video nào cho từ khóa '{keyword}'. Bạn có thể bấm 'Mở Trình duyệt Chrome' để tự giải Captcha nếu TikTok yêu cầu!")

    def clear_results(self):
        for row in self.tree.get_children():
            self.tree.delete(row)
        self.data_results = []
        self.lbl_count.config(text="Tổng số: 0 video")
        self.lbl_status.config(text="Đã xóa danh sách.", fg="#555555")
        self.btn_export.config(state="disabled")

    def export_to_excel(self):
        if not self.data_results:
            messagebox.showwarning("Cảnh báo", "Chưa có dữ liệu để xuất Excel!")
            return

        keyword = self.txt_keyword.get().strip().replace(" ", "_")
        time_str = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"tiktok_{keyword}_{time_str}.xlsx"

        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel Files (*.xlsx)", "*.xlsx"), ("CSV Files (*.csv)", "*.csv")],
            initialfile=default_filename,
            title="Lưu file Excel"
        )

        if not file_path:
            return

        try:
            if file_path.endswith(".csv") or not HAS_OPENPYXL:
                import csv
                with open(file_path, "w", newline="", encoding="utf-8-sig") as f:
                    writer = csv.writer(f)
                    writer.writerow(["Link", "Caption"])
                    for item in self.data_results:
                        writer.writerow([item["url"], item["caption"]])
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "TikTok Videos"

                header_fill = PatternFill(start_color="FE2C55", end_color="FE2C55", fill_type="solid")
                header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
                center_align = Alignment(horizontal="center", vertical="center")
                left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )

                # Write Headers: Link & Caption
                headers = ["Link", "Caption"]
                ws.append(headers)

                ws.row_dimensions[1].height = 28
                for col_idx in range(1, 3):
                    cell = ws.cell(row=1, column=col_idx)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_align

                # Write Data rows
                for row_idx, item in enumerate(self.data_results, start=2):
                    ws.append([item["url"], item["caption"]])
                    ws.row_dimensions[row_idx].height = 24
                    
                    link_cell = ws.cell(row=row_idx, column=1)
                    link_cell.font = Font(name="Segoe UI", size=10, color="0044CC", underline="single")
                    link_cell.alignment = left_align
                    link_cell.border = thin_border
                    
                    cap_cell = ws.cell(row=row_idx, column=2)
                    cap_cell.font = Font(name="Segoe UI", size=10)
                    cap_cell.alignment = left_align
                    cap_cell.border = thin_border

                ws.column_dimensions['A'].width = 50
                ws.column_dimensions['B'].width = 80

                wb.save(file_path)

            self.lbl_status.config(text=f"✅ Đã xuất file thành công: {os.path.basename(file_path)}", fg="#198754")
            
            open_now = messagebox.askyesno(
                "Thành công", 
                f"Đã xuất thành công {len(self.data_results)} video thật vào file:\n{file_path}\n\nBạn có muốn mở file này ngay không?"
            )
            if open_now:
                os.startfile(file_path)

        except Exception as e:
            messagebox.showerror("Lỗi xuất file", f"Không thể lưu file Excel:\n{str(e)}")

def main():
    root = tk.Tk()
    app = TikTokSearchApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
