from pathlib import Path
from typing import Optional, Dict, Any, List
from fastapi import FastAPI, HTTPException, Body, UploadFile, File
from fastapi.responses import HTMLResponse, JSONResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel

from tollcal.domain.states import RightsBasis
from tollcal.media.batch_parser import generate_sample_template, parse_batch_file
from tollcal.observability.logging import logger
from tollcal.storage.database import init_database
from tollcal.storage.repositories import ChannelRepository, JobRepository
from tollcal.sync.orchestrator import SyncOrchestrator
from tollcal.ucircle.client import UCircleClient
from tollcal.config import settings

app = FastAPI(title="TOLLCAL Web API", version="1.0.0")


@app.get("/api/template/batch-excel")
def get_batch_template():
    """Tải file Excel mẫu chứa danh sách link TikTok."""
    file_path = generate_sample_template()
    return FileResponse(
        path=str(file_path.resolve()),
        filename="mau_danh_sach_tiktok.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.post("/api/sync/parse-batch-file")
async def parse_batch_file_endpoint(file: UploadFile = File(...)):
    """Tải lên file Excel (.xlsx) hoặc CSV để trích xuất danh sách link TikTok cần đồng bộ."""
    try:
        content = await file.read()
        items = parse_batch_file(content, file.filename)
        return {"ok": True, "count": len(items), "items": items}
    except Exception as e:
        logger.error(f"[Batch Parse] Lỗi phân tích file: {e}")
        return {"ok": False, "error": f"Lỗi đọc file: {str(e)}"}

# Enable CORS for local access
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# Khởi tạo DB
init_database()
orchestrator = SyncOrchestrator()
HTML_TEMPLATE_PATH = Path(__file__).parent / "templates" / "index.html"


class SendOtpRequest(BaseModel):
    email: str


class VerifyOtpRequest(BaseModel):
    email: str
    token: str


class SyncUrlRequest(BaseModel):
    url: str
    visibility: str = "public"
    circle_id: Optional[str] = None
    caption: Optional[str] = None


class AddChannelRequest(BaseModel):
    url: str
    creator_id: str
    rights: str = "owner"


@app.get("/", response_class=HTMLResponse)
def get_index():
    """Phục vụ trang giao diện Web UI."""
    if HTML_TEMPLATE_PATH.exists():
        return HTMLResponse(content=HTML_TEMPLATE_PATH.read_text(encoding="utf-8"))
    return HTMLResponse(content="<h1>UI Template not found</h1>", status_code=404)


@app.get("/api/auth/status")
def get_auth_status():
    """Kiểm tra trạng thái đăng nhập và thông tin hạn mức."""
    client = UCircleClient()
    try:
        session = client.auth.login()
        caps = client.get_caps()
        return {
            "authenticated": True,
            "user_id": session.user_id,
            "email": session.email,
            "caps": caps.model_dump(),
            "default_circle_id": settings.default_circle_id,
        }
    except Exception as e:
        return {
            "authenticated": False,
            "error": str(e),
            "user_id": None,
            "caps": None,
        }


@app.get("/api/circles")
def list_user_circles_endpoint():
    """Lấy danh sách Circles của tài khoản hiện tại."""
    try:
        client = UCircleClient()
        circles = client.list_my_circles()
        return {"ok": True, "circles": circles, "default_circle_id": settings.default_circle_id}
    except Exception as e:
        return {"ok": False, "error": str(e), "circles": [], "default_circle_id": settings.default_circle_id}


@app.post("/api/auth/send-otp")
def send_otp_endpoint(req: SendOtpRequest):
    """Gửi mã OTP về email."""
    client = UCircleClient()
    try:
        client.auth.send_otp(req.email)
        return {"ok": True, "message": f"Đã gửi mã OTP tới {req.email}"}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/auth/verify-otp")
def verify_otp_endpoint(req: VerifyOtpRequest):
    """Xác thực mã OTP và lưu phiên đăng nhập."""
    client = UCircleClient()
    try:
        session = client.auth.verify_otp(req.email, req.token)
        return {"ok": True, "user_id": session.user_id, "email": session.email}
    except Exception as e:
        return {"ok": False, "error": str(e)}


@app.post("/api/sync/url")
def sync_url_endpoint(req: SyncUrlRequest):
    """Thực thi đồng bộ 1 video TikTok sang UCircle Wavee."""
    try:
        orch = SyncOrchestrator()
        job = orch.sync_single_video(
            source_url_or_path=req.url,
            visibility=req.visibility,
            circle_id=req.circle_id,
            custom_caption=req.caption,
            poll_encode=True,
        )
        if not job:
            raise RuntimeError("Đồng bộ hoàn tất nhưng không lấy được bản ghi Job.")
        return {"ok": True, "job": job.model_dump()}
    except Exception as e:
        logger.error(f"[API Sync] Lỗi: {e}")
        return {"ok": False, "error": str(e)}


@app.get("/api/export/csv")
@app.get("/api/export/excel")
def download_csv_endpoint():
    """Xuất và tải file CSV lịch sử đồng bộ theo ngày hiện tại."""
    from tollcal.storage.excel_export import export_jobs_to_csv
    
    file_path = export_jobs_to_csv()
    if file_path.exists():
        return FileResponse(
            path=str(file_path.resolve()),
            filename=file_path.name,
            media_type="text/csv",
        )
    raise HTTPException(status_code=404, detail="Chưa có dữ liệu lịch sử để xuất.")


@app.get("/api/jobs")
def list_jobs_endpoint():
    """Lấy danh sách các jobs gần nhất."""
    jobs = JobRepository.list_jobs(limit=50)
    return [j.model_dump() for j in jobs]


@app.get("/api/channels")
def list_channels_endpoint():
    """Lấy danh sách các kênh đang theo dõi."""
    channels = ChannelRepository.list_active_channels()
    return [ch.model_dump() for ch in channels]


@app.post("/api/channels")
def add_channel_endpoint(req: AddChannelRequest):
    """Thêm một kênh mới vào danh sách theo dõi."""
    rights_basis = RightsBasis(req.rights) if req.rights in [e.value for e in RightsBasis] else RightsBasis.OWNER
    try:
        ch = ChannelRepository.add_channel(req.url, req.creator_id, rights_basis)
        return {"ok": True, "channel": ch.model_dump()}
    except Exception as e:
        return {"ok": False, "error": str(e)}


class ScanChannelRequest(BaseModel):
    channel_url: str
    mode: str = "auto"  # "auto", "search", "channel"
    max_videos: int = 100
    method: str = "auto"
    max_duration: Optional[float] = 120.0  # Mặc định lọc video <= 120s (2 phút)


@app.post("/api/tiktok/scan-channel")
def scan_channel_endpoint(req: ScanChannelRequest):
    """Quét toàn bộ video từ một kênh hoặc Tìm kiếm video theo Từ khóa TikTok."""
    from tollcal.sources.tiktok_channel_scanner import channel_scanner
    
    query = req.channel_url.strip()
    if not query:
        return {"ok": False, "error": "Vui lòng nhập từ khóa tìm kiếm hoặc link kênh TikTok."}

    try:
        # Xác định mode
        is_search = False
        if req.mode == "search":
            is_search = True
        elif req.mode == "channel":
            is_search = False
        else:
            # Auto: nếu không chứa tiktok.com/@ hoặc không bắt đầu bằng @ -> tìm kiếm
            if not query.startswith("@") and "tiktok.com/@" not in query and not query.startswith("http"):
                is_search = True

        if is_search:
            videos = channel_scanner.search(
                keywords=query,
                max_videos=req.max_videos,
                method=req.method,
                max_duration=req.max_duration,
            )
            return {
                "ok": True,
                "mode": "search",
                "query": query,
                "count": len(videos),
                "videos": videos,
            }
        else:
            username = channel_scanner.extract_username(query)
            videos = channel_scanner.scan(
                channel_input=query,
                max_videos=req.max_videos,
                method=req.method,
                max_duration=req.max_duration,
            )
            return {
                "ok": True,
                "mode": "channel",
                "username": username,
                "count": len(videos),
                "videos": videos,
            }
    except Exception as e:
        logger.error(f"[API Scan Channel / Search] Lỗi: {e}")
        return {"ok": False, "error": str(e)}


class ExportScannedRequest(BaseModel):
    videos: List[Dict[str, Any]]
    circle_id: Optional[str] = None
    visibility: str = "public"


@app.post("/api/tiktok/export-scanned-excel")
def export_scanned_excel_endpoint(req: ExportScannedRequest):
    """Xuất danh sách video vừa quét được thành file Excel (.xlsx) chuẩn."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    import tempfile

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Videos Quét Được"

    headers = [
        "Link TikTok (Bắt buộc)",
        "Caption Tùy Chỉnh (Tùy chọn)",
        "Chế Độ Hiển Thị (public / connections)",
        "Circle ID (Tùy chọn)",
    ]
    ws.append(headers)

    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for v in req.videos:
        ws.append([
            v.get("url", ""),
            v.get("title", ""),
            req.visibility or "public",
            req.circle_id or "",
        ])

    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    wb.save(temp_file.name)
    temp_file.close()

    return FileResponse(
        path=temp_file.name,
        filename="danh_sach_video_tiktok_quet_duoc.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
