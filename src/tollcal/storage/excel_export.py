import csv
from datetime import datetime
from pathlib import Path
from typing import List, Optional

from tollcal.config import settings
from tollcal.domain.models import SyncJob
from tollcal.observability.logging import logger
from tollcal.storage.repositories import JobRepository


def get_daily_file_paths() -> tuple[Path, Path]:
    """Lấy đường dẫn file Excel (.xlsx) và CSV (.csv) tự động theo ngày hiện tại."""
    today_str = datetime.now().strftime("%Y-%m-%d")
    raw_path_str = str(settings.csv_export_path)

    base_stem = f"lich_su_{today_str}"
    if "{date}" in raw_path_str:
        base_stem = raw_path_str.replace("{date}", today_str)
    elif "%Y" in raw_path_str:
        base_stem = datetime.now().strftime(raw_path_str)

    base_path = Path(base_stem)
    parent_dir = base_path.parent
    stem = base_path.stem
    if stem.endswith(f"_{today_str}"):
        stem = base_path.stem
    elif not stem.endswith(today_str):
        stem = f"{stem}_{today_str}" if stem != "lich_su" else f"lich_su_{today_str}"

    parent_dir.mkdir(parents=True, exist_ok=True)
    excel_path = parent_dir / f"{stem}.xlsx"
    csv_path = parent_dir / f"{stem}.csv"
    return excel_path, csv_path


def get_ucircle_watch_url(video_id: Optional[str]) -> str:
    """Tạo đường link xem video trực tiếp trên UCircle Wavee."""
    if not video_id or video_id == "—":
        return "—"
    base = settings.ucircle_base_url.rstrip("/")
    return f"{base}/wavee/{video_id}"


def export_jobs_to_files(custom_excel_path: Optional[Path] = None, custom_csv_path: Optional[Path] = None) -> tuple[Path, Path]:
    """
    Xuất lịch sử đồng bộ sang cả 2 định dạng:
    1. File Excel (.xlsx): TỰ ĐỘNG CÂN CHỈNH ĐỘ RỘNG CỘT theo nội dung, có màu sắc, viền bảng và link bấm trực tiếp.
    2. File CSV (.csv): Chuẩn utf-8-sig để xem nhanh trên mọi thiết bị.
    """
    default_excel, default_csv = get_daily_file_paths()
    target_excel = custom_excel_path or default_excel
    target_csv = custom_csv_path or default_csv

    target_excel.parent.mkdir(parents=True, exist_ok=True)
    target_csv.parent.mkdir(parents=True, exist_ok=True)

    jobs: List[SyncJob] = JobRepository.list_jobs(limit=2000)

    # 1. Ghi file CSV
    try:
        with open(target_csv, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow([
                "STT",
                "Nguồn",
                "Mã Video TikTok",
                "Link Video TikTok",
                "Mã Video UCircle",
                "Link Video UCircle Wavee",
                "Tiêu đề / Caption",
                "Thời lượng (giây)",
                "Dung lượng (MB)",
                "Trạng thái",
                "Thời gian đồng bộ",
                "Ghi chú / Lỗi",
            ])
            for idx, j in enumerate(jobs, 1):
                size_mb = f"{j.file_size_bytes / (1024 * 1024):.2f}" if j.file_size_bytes else "0.0"
                ucircle_link = get_ucircle_watch_url(j.ucircle_video_id)
                writer.writerow([
                    idx,
                    j.source_provider.upper(),
                    j.source_video_id,
                    j.source_url,
                    j.ucircle_video_id or "—",
                    ucircle_link,
                    j.caption,
                    f"{j.duration_seconds:.1f}" if j.duration_seconds else "0",
                    size_mb,
                    j.state.value,
                    str(j.created_at)[:19],
                    j.error_message or "",
                ])
    except Exception as e:
        logger.warning(f"[Export] Lỗi ghi file CSV: {e}")

    # 2. Ghi file Excel (.xlsx) với TỰ ĐỘNG ĐIỀU CHỈNH KÍCH THƯỚC CỘT & ĐỊNH DẠNG ĐẸP
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Lịch Sử Đồng Bộ"

        headers = [
            "STT",
            "Nguồn",
            "Mã Video TikTok",
            "Link Video TikTok",
            "Mã Video UCircle",
            "Link Video UCircle Wavee",
            "Tiêu đề / Caption",
            "Thời lượng (s)",
            "Dung lượng (MB)",
            "Trạng thái",
            "Thời gian tạo",
            "Ghi chú / Lỗi",
        ]
        ws.append(headers)

        # Style Header (Màu xanh Navy cao cấp)
        header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)

        for col_num, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align

        green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        yellow_fill = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='E2E8F0'),
            right=Side(style='thin', color='E2E8F0'),
            top=Side(style='thin', color='E2E8F0'),
            bottom=Side(style='thin', color='E2E8F0'),
        )

        for row_idx, j in enumerate(jobs, 2):
            size_mb = round(j.file_size_bytes / (1024 * 1024), 2) if j.file_size_bytes else 0.0
            ucircle_link = get_ucircle_watch_url(j.ucircle_video_id)

            row_data = [
                row_idx - 1,
                j.source_provider.upper(),
                j.source_video_id,
                j.source_url,
                j.ucircle_video_id or "—",
                ucircle_link,
                j.caption,
                round(j.duration_seconds, 1) if j.duration_seconds else 0.0,
                size_mb,
                j.state.value,
                str(j.created_at)[:19],
                j.error_message or "",
            ]
            ws.append(row_data)

            # Format cell
            for col_idx in range(1, len(headers) + 1):
                c = ws.cell(row=row_idx, column=col_idx)
                c.font = Font(name="Segoe UI", size=10)
                c.border = thin_border
                c.alignment = Alignment(vertical="center")

                # Căn giữa các cột ngắn
                if col_idx in (1, 2, 8, 9, 10, 11):
                    c.alignment = center_align

                # Biến link thành Hyperlink có màu xanh nhạt
                if col_idx in (4, 6) and str(c.value).startswith("http"):
                    c.font = Font(name="Segoe UI", size=10, color="0284C7", underline="single")

                # Tô màu trạng thái
                if col_idx == 10:
                    if j.state.value in ("PUBLISHED", "ENCODED"):
                        c.fill = green_fill
                        c.font = Font(name="Segoe UI", size=10, bold=True, color="166534")
                    elif "FAIL" in j.state.value or "BLOCKED" in j.state.value:
                        c.fill = red_fill
                        c.font = Font(name="Segoe UI", size=10, bold=True, color="991B1B")
                    else:
                        c.fill = yellow_fill
                        c.font = Font(name="Segoe UI", size=10, bold=True, color="854D0E")

        # TỰ ĐỘNG CÂN CHỈNH ĐỘ RỘNG CỘT BẰNG ĐÚNG KÍCH THƯỚC NỘI DUNG
        for col in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col[0].column)
            for cell in col:
                val = str(cell.value or "")
                # Tính độ dài thực tế của text
                max_len = max(max_len, len(val))
            
            # Gán độ rộng cột = độ dài dài nhất + 4 ký tự padding (tối thiểu 10, tối đa 65)
            ws.column_dimensions[col_letter].width = min(max(max_len + 4, 10), 65)

        wb.save(target_excel)
        logger.debug(f"[Export] Đã xuất file Excel tự co giãn cột: {target_excel.resolve()}")
    except ImportError:
        pass
    except Exception as e:
        logger.warning(f"[Export] Lỗi tạo file Excel: {e}")

    return target_excel, target_csv


def export_jobs_to_csv(custom_path: Optional[Path] = None) -> Path:
    _, csv_path = export_jobs_to_files(custom_csv_path=custom_path)
    return csv_path


def export_jobs_to_excel(excel_path: Optional[Path] = None, csv_path: Optional[Path] = None) -> Path:
    excel_p, _ = export_jobs_to_files(custom_excel_path=excel_path, custom_csv_path=csv_path)
    return excel_p
