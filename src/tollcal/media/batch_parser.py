import csv
import io
import re
from pathlib import Path
from typing import List, Dict, Any, Optional
from tollcal.observability.logging import logger

SAMPLE_TEMPLATE_PATH = Path("mau_danh_sach_tiktok.xlsx")


def generate_sample_template() -> Path:
    """Tạo file Excel mẫu để người dùng điền danh sách link TikTok."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Danh Sách TikTok"

    headers = [
        "Link TikTok (Bắt buộc)",
        "Caption Tùy Chỉnh (Tùy chọn)",
        "Chế Độ Hiển Thị (public / connections)",
        "Circle ID (Tùy chọn)",
    ]
    ws.append(headers)

    # Style Header
    header_fill = PatternFill(start_color="0284C7", end_color="0284C7", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Mẫu dữ liệu ví dụ
    samples = [
        ["https://www.tiktok.com/@creator1/video/7123456789012345678", "Video hài hước xu hướng #trend #haihuoc", "public", ""],
        ["https://vt.tiktok.com/ZSjSample1/", "Tổng hợp mẹo hay cuộc sống", "public", ""],
        ["https://vt.tiktok.com/ZSjSample2/", "", "connections", ""],
    ]
    for row in samples:
        ws.append(row)

    # Auto width
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 15)

    wb.save(SAMPLE_TEMPLATE_PATH)
    return SAMPLE_TEMPLATE_PATH


def extract_circle_id(raw_val: Any) -> Optional[str]:
    """
    Trích xuất UUID của Circle một cách an toàn.
    Hỗ trợ cả trường hợp người dùng dán cả URL (ví dụ https://ucircle.net/ci/d159b031-169d-4b4d-a2af-e8087b84a1fb)
    hoặc chỉ dán chuỗi UUID (d159b031-169d-4b4d-a2af-e8087b84a1fb).
    """
    if not raw_val:
        return None
    val_str = str(raw_val).strip()
    if not val_str or val_str.lower() in ("none", "null", "trống", ""):
        return None
    
    # Tìm chuỗi UUID 36 ký tự
    match = re.search(r'[0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12}', val_str)
    if match:
        return match.group(0).lower()
    return val_str


def parse_batch_file(file_content: bytes, filename: str) -> List[Dict[str, Any]]:
    """
    Phân tích file Excel (.xlsx) hoặc CSV để trích xuất danh sách link TikTok và caption.
    Tự động nhận diện cột URL thông minh (theo tên cột hoặc tìm link tiktok trong ô).
    """
    items = []
    lower_fn = filename.lower()

    if lower_fn.endswith(".xlsx") or lower_fn.endswith(".xls"):
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file_content), data_only=True)
        ws = wb.active

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            return []

        # Tìm vị trí các cột
        header_row = [str(c or "").strip().lower() for c in rows[0]]
        url_idx = 0
        caption_idx = None
        vis_idx = None
        circle_idx = None

        for idx, h in enumerate(header_row):
            if any(k in h for k in ["link", "url", "tiktok", "video"]):
                url_idx = idx
            elif any(k in h for k in ["caption", "title", "tiêu đề", "mô tả"]):
                caption_idx = idx
            elif any(k in h for k in ["visibility", "chế độ", "hiển thị"]):
                vis_idx = idx
            elif any(k in h for k in ["circle", "nhóm", "page", "trang"]):
                circle_idx = idx

        # Nếu dòng đầu là tiêu đề thì duyệt từ dòng 2, ngược lại duyệt từ dòng 1
        has_header = any("link" in h or "url" in h or "tiktok" in h for h in header_row)
        start_row = 1 if has_header else 0

        for r in rows[start_row:]:
            if not r or not any(r):
                continue
            # Tìm link tiktok trong row
            url = None
            if url_idx < len(r) and r[url_idx] and "tiktok.com" in str(r[url_idx]):
                url = str(r[url_idx]).strip()
            else:
                for val in r:
                    if val and "tiktok.com" in str(val):
                        url = str(val).strip()
                        break

            if url:
                caption = str(r[caption_idx]).strip() if (caption_idx is not None and caption_idx < len(r) and r[caption_idx]) else None
                visibility = str(r[vis_idx]).strip().lower() if (vis_idx is not None and vis_idx < len(r) and r[vis_idx]) else "public"
                raw_circle = r[circle_idx] if (circle_idx is not None and circle_idx < len(r)) else None
                circle_id = extract_circle_id(raw_circle)

                items.append({
                    "url": url,
                    "caption": caption,
                    "visibility": "connections" if visibility == "connections" else "public",
                    "circle_id": circle_id,
                })

    else:
        # Xử lý file CSV
        text = file_content.decode("utf-8-sig", errors="ignore")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        if not rows:
            return []

        header_row = [c.strip().lower() for c in rows[0]]
        url_idx = 0
        caption_idx = None
        vis_idx = None
        circle_idx = None

        for idx, h in enumerate(header_row):
            if any(k in h for k in ["link", "url", "tiktok", "video"]):
                url_idx = idx
            elif any(k in h for k in ["caption", "title", "tiêu đề"]):
                caption_idx = idx
            elif any(k in h for k in ["visibility", "chế độ"]):
                vis_idx = idx
            elif any(k in h for k in ["circle", "nhóm", "page", "trang"]):
                circle_idx = idx

        has_header = any("link" in h or "url" in h or "tiktok" in h for h in header_row)
        start_row = 1 if has_header else 0

        for r in rows[start_row:]:
            if not r:
                continue
            url = None
            if url_idx < len(r) and "tiktok.com" in r[url_idx]:
                url = r[url_idx].strip()
            else:
                for val in r:
                    if "tiktok.com" in val:
                        url = val.strip()
                        break

            if url:
                caption = r[caption_idx].strip() if (caption_idx is not None and caption_idx < len(r) and r[caption_idx]) else None
                visibility = r[vis_idx].strip().lower() if (vis_idx is not None and vis_idx < len(r) and r[vis_idx]) else "public"
                raw_circle = r[circle_idx] if (circle_idx is not None and circle_idx < len(r)) else None
                circle_id = extract_circle_id(raw_circle)

                items.append({
                    "url": url,
                    "caption": caption,
                    "visibility": "connections" if visibility == "connections" else "public",
                    "circle_id": circle_id,
                })

    return items
